#!/usr/bin/env python3
import os
import argparse
import logging
import sys
import logging
from rich.logging import RichHandler
import configparser
import openpyxl
from typing import NamedTuple

import json
import paho.mqtt.client as mqtt

#own
import utils_influx_Client
import export
import util

LOG_LEVELS = ["DEBUG", "INFO", "WARNING", "ERROR", "CRITICAL"]
DEFAULT_LOG_LEVEL = "INFO"

class SensorData(NamedTuple):
    friendlyGW: str
    eepId: str
    deviceId: int
    label: str
    label_short: str
    dbm: int
    sensortype: str
    room: str
    value: int
    floor: int    
    voltage: float    
    count: int    
    timestamp: int  

 

def on_connect(client, userdata, flags, rc):
    """ The callback for when the client receives a CONNACK response from the server."""
    log.info('Connected to MQTT Broker with result code ' + str(rc))
    client.subscribe(config['MQTT_TOPIC'])

def on_message(client, userdata, msg):
    """The callback for when a PUBLISH message is received from the server."""
    sensor_data = _parse_mqtt_message(msg.payload.decode('utf-8'))
    if sensor_data is not None:
        send_sensor_data_to_influxdb(sensor_data)


def _parse_mqtt_message(payload):
    log.debug(payload)
    data = json.loads(payload)

    friendlyGW = data['friendlyGW']
    eepId = data['telegram']['eepId']
    deviceId = data['telegram']['deviceId']
    label = data['telegram']['friendlyId']

    floor,room,type,count = label.split('-')
    floor = floor.replace('O','')
    floor = floor.replace('E','')
    label_short = room + "-" + count 

    dbm = data['telegram']['dbm']
    timestamp = data['telegram']['timestamp']

    sensortype = json.dumps(data['telegram']['functions'][0]['key']).replace(' ','_')   
    sensortype = sensortype.replace('"','')   

    value = json.dumps(data['telegram']['functions'][0]['value'])    
    voltage = json.dumps(data['telegram']['functions'][1]['value'])

    return SensorData(friendlyGW, eepId, deviceId, label, label_short, int(dbm), sensortype, room, int(value), int(floor), float(voltage), int(count), timestamp)


def send_sensor_data_to_influxdb(sensor_data):
    log.debug(sensor_data)
    json_body = [
        {
            'measurement': sensor_data.sensortype,
            'tags': {
                'gateway': sensor_data.friendlyGW,
                'type': sensor_data.sensortype,
                'eepId': sensor_data.eepId,
                'deviceId': sensor_data.deviceId,
                'label': sensor_data.label,
                'label_short': sensor_data.label_short,                
                'floor': sensor_data.floor,
                'room': sensor_data.room,
                'deviceId': sensor_data.deviceId
            },
            'fields': {                
                'count': sensor_data.count,
                'value': sensor_data.value,
                'voltage': sensor_data.voltage,
                'dbm': sensor_data.dbm,
                'timestamp': sensor_data.timestamp
            }
        }
    ]
    log.debug(json_body)
    log.debug(influxdb_client.write_points(json_body,time_precision="ms",protocol = "json"))

def mqtt_loop():
    mqtt_client = mqtt.Client(config['MQTT_CLIENT_ID'])
    mqtt_client.username_pw_set(config['MQTT_USER'], config['MQTT_PASSWORD'])
    mqtt_client.on_connect = on_connect
    mqtt_client.on_message = on_message
    mqtt_client.connect(config['MQTT_ADDRESS'], int(config['MQTT_PORT']) )
    mqtt_client.loop_forever()


    
def runExcelReader(xlsFile, output_path, svg_path, pdf_path, sheet="Vossloh und Schwabe Multisensor"):  

    if not os.path.exists(xlsFile):
        log.error("[bold red blink]Excel File not found " + xlsFile + " [/]", extra={"markup": True})          
        log.error("Exit")  


    wb_obj = openpyxl.load_workbook(xlsFile)
    sheet_obj = wb_obj[sheet]
    m_row = sheet_obj.max_row
    
    # Loop will print all values
    for i in range(2, m_row + 1):
        eshellid_obj = sheet_obj.cell(row = i, column = 1)
        label_obj = sheet_obj.cell(row = i, column = 2)
        gw_obj = sheet_obj.cell(row = i, column = 3)
        floor_obj = sheet_obj.cell(row = i, column = 4)
        room_obj = sheet_obj.cell(row = i, column = 5)
        id_obj = sheet_obj.cell(row = i, column = 6)
        qr_obj = sheet_obj.cell(row = i, column = 7)        

    #     try:
    #         if len(gw_obj.value) > 0: 
    #             svg_path2 = export.createFolderPath(svg_path, gw_obj.value)
    #             pdf_path2 = export.createFolderPath(pdf_path, gw_obj.value)
    #             svg_filename_label = os.path.join(svg_path2, label_obj.value + "_" + id_obj.value + ".svg")
    #             pdf_filename_label = os.path.join(pdf_path2, label_obj.value + "_" + id_obj.value + ".pdf")
    #             log.debug("svg_filename: " + svg_filename_label)
    #             log.debug("pdf_filename: " + pdf_filename_label)
    #     except:
    #         pass    

    #     try:
    #         if len(gw_obj.value) > 0:
    #             log.info("eshellid: " + str(eshellid_obj.value) + " | label: " + str(label_obj.value) + " | gw: " + str(gw_obj.value) + 
    #                 " | floor: " + str(floor_obj.value) + " | room: " + str(room_obj.value) + " | id: " + str(id_obj.value) + 
    #                 " | qr: " + str(qr_obj.value)  )   
    #         createQR.make_pdf_file(svg_filename_label, pdf_filename_label, str(qr_obj.value), str(floor_obj.value), str(room_obj.value), str(label_obj.value), str(gw_obj.value), str(eshellid_obj.value), str(id_obj.value))  
    #     except:
    #         pass    

    # #merge all PDF to one PDF per Gateway        
    # mergePDFs(pdf_path,"pdf",output_path)
    # log.info("[bold green blink]- Finished! " + str(xlsFile) + " [/]", extra={"markup": True})



def allFilesMode(path,fileextension):
    #Reads all files from Folder          
    for filename in os.listdir(path):
        if filename.endswith(fileextension) :        
            log.info("Process File: " + os.path.join(path, filename))
            runExcelReader(os.path.join(path,filename), "Vossloh und Schwabe Multisensor")
        else:
            log.info("Finished. No more Files found in:" + path)
            continue  

def get_parser():
    """Get parser object """
    from argparse import ArgumentParser, ArgumentDefaultsHelpFormatter
    parser = argparse.ArgumentParser(description='Understand functioning')
    parser.add_argument("-x", "--xlsx", type=str, required=False,  metavar="FILE", help="Path to Excel File", default="") 
    parser.add_argument("-o", "--outputfolder", type=str, required=False,  metavar="FILE", help="Path to output folder", default="") 
    parser.add_argument("-s", "--xlsfolder", type=str, required=False,  metavar="FILE", help="Path to Excel Files Folder", default="") 
    parser.add_argument("-l", "--logfile", type=str, required=False, help="Logfile", default="")
    parser.add_argument("-c", "--config", type=str, required=False, help="Config", default="")
    parser.add_argument("--verbose", "-v", dest="log_level", action="append_const", const=-1,)
    parser.add_argument("--quiet", "-q", dest="log_level", action="append_const", const=1,)
    return parser

def setLogging(args):   
    log_level = LOG_LEVELS.index(DEFAULT_LOG_LEVEL)
    for adjustment in args.log_level or ():
            log_level = min(len(LOG_LEVELS) - 1, max(log_level + adjustment, 0))

    log_level_name = LOG_LEVELS[log_level]

    global log
    log = logging.getLogger(__name__)    
    shell_handler = RichHandler()    
    log.setLevel(log_level_name)
    shell_handler.setLevel(log_level_name)    
    fmt_shell = '%(message)s'    
    shell_formatter = logging.Formatter(fmt_shell)
    shell_handler.setFormatter(shell_formatter)
    log.addHandler(shell_handler)
    
    if len(args.logfile) > 0:
        file_handler = logging.FileHandler(args.logfile)
        file_handler.setLevel(logging.DEBUG)
        fmt_file = '%(levelname)s %(asctime)s [%(filename)s:%(funcName)s:%(lineno)d] %(message)s'
        file_formatter = logging.Formatter(fmt_file)
        file_handler.setFormatter(file_formatter)
        log.addHandler(file_handler)        

def main():

    log.info("[bold green]MQTT2Influx EnOcean © Stefan Knaak - E-Shelter Security.  [/]", extra={"markup": True}) 
    
    #Folders
    cwd = os.getcwd()
    xls_path = export.checkInputFolderPath(args.xlsfolder, "xlsx")
    log.debug("xls_path: " + str(xls_path))    
    
    #Operation Mode
    #Get Excel File from Command line
    if len(args.xlsx) > 0 :
        if os.path.isfile(args.xlsx) :
            log.info("[bold purple]Single File Mode.  [/]", extra={"markup": True})  
            log.info("[bold green]Image: " + args.xlsx + " [/]", extra={"markup": True})  
            #runExcelReader(args.xlsx,"Vossloh und Schwabe Multisensor")
        else:  
            log.info("[bold cyan]Single File Mode.  [/]", extra={"markup": True})  
            if not os.path.isfile(args.xlsx):
                log.error("[bold red]Excel File not found [/] [yellow] " + args.xlsx + " [/]", extra={"markup": True})  
            log.error("Exit")    

    else: #/data/xlsx folder
        log.info("[bold cyan]Folder Mode[/]", extra={"markup": True}) 
        if len(args.xlsfolder) > 0:
            log.info("[bold white]Folder: [/] " + xls_path , extra={"markup": True}) 

        if not os.path.exists(xls_path):
            log.error("[bold red blink]xlsx Folder not found " + xls_path + " [/]", extra={"markup": True})  
            log.error("Exit")  
        else:  
            log.info("[bold cyan]Process all Excel Files from " + xls_path + " [/]", extra={"markup": True})     
            #allFilesMode(xls_path,'xlsx')       

    #Main Loop
    mqtt_loop()        

def config_read():
    config_f = configparser.ConfigParser()

    try:
        config_f.read_file(open(args.config))
    
        config['MQTT_ADDRESS']      = config_f.get('mqtt', 'MQTT_ADDRESS')
        config['MQTT_PORT']         = config_f.get('mqtt', 'MQTT_PORT')
        config['MQTT_USER']         = config_f.get('mqtt', 'MQTT_USER')
        config['MQTT_PASSWORD']     = config_f.get('mqtt', 'MQTT_PASSWORD')
        config['MQTT_TOPIC']        = config_f.get('mqtt', 'MQTT_TOPIC')
        config['MQTT_CLIENT_ID']    = config_f.get('mqtt', 'MQTT_CLIENT_ID')

        config['INFLUXDB_ADDRESS']  = config_f.get('influx', 'INFLUXDB_ADDRESS')
        config['INFLUXDB_PORT']     = config_f.get('influx', 'INFLUXDB_PORT')
        config['INFLUXDB_USER']     = config_f.get('influx', 'INFLUXDB_USER')
        config['INFLUXDB_PASSWORD'] = config_f.get('influx', 'INFLUXDB_PASSWORD')
        config['INFLUXDB_DATABASE'] = config_f.get('influx', 'INFLUXDB_DATABASE')
        config['INFLUXDB_USESLL']   = config_f.get('influx', 'INFLUXDB_USESLL')
        config['INFLUXDB_NOVERIFY'] = config_f.get('influx', 'INFLUXDB_NOVERIFY')
        config['INFLUXDB_CREATE']   = config_f.get('influx', 'INFLUXDB_CREATE')
    
    except configparser.ParsingError:
        log.error("Unable to parse config from file %s!", args.config)
        sys.exit(1)

              
if __name__ == "__main__":    

    #parse arguments
    args = get_parser().parse_args()

    #get config
    if not args.config:
        args.config = os.path.dirname(os.path.abspath(__file__)) + '/config_enocean.conf'

    #get config data
    config = {}
    config_read()

    setLogging(args)

    #create influx connection
    influxdb_client = utils_influx_Client.influxClient(config,log)

    try:
        main()
    except BaseException:
        import sys
        print(sys.exc_info()[0])
        import traceback
        print(traceback.format_exc())
    # finally:
    #     print("Press Enter to continue ...")
    #     input()
