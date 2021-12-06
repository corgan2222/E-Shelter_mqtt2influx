#!/usr/bin/env python3
import os
import argparse
import logging
import sys
import logging
from rich.logging import RichHandler
import configparser
import openpyxl
from typing import NamedTuple
import csv
from csv import DictReader
import json
import paho.mqtt.client as mqtt

#own
import utils_influx_Client
import export
import util

LOG_LEVELS = ["DEBUG", "INFO", "WARNING", "ERROR", "CRITICAL"]
DEFAULT_LOG_LEVEL = "INFO"

#global
csvdata = {}
sheet_obj = {}

class SensorData(NamedTuple):
    deviceUuid: str
    deviceId: str
    siteUuid: str
    type: str
    label: str
    label_short: str
    gateway: str
    room: str
    floor: int
    index: int    
    value: int    
    timestamp: int  


def _parse_mqtt_message(payload):
    #log.debug(payload)
    data = json.loads(payload)

    deviceUuid = data['deviceUuid']
    deviceId = data['deviceId']
    siteUuid = data['siteUuid']
    type = data['type']
    index = data['index']
    value = data['value']
    timestamp = data['timestamp']

    log.debug("did mqtt " + deviceId)

    #Thing IT Export CSV
    # for row in csvdata:
    #       print(row)

    #AssetListe
    label_short = ""
    room = ""
    label = ""
    gateway = ""
    floor = -1

    # Loop will print all values
    m_row = sheet_obj.max_row
    for i in range(2, m_row + 1):
        id_obj = sheet_obj.cell(row = i, column = 6)                
        
        try:
            if deviceId == str(id_obj.value):
                log.debug("did excel " + id_obj.value)

                eshellid_obj = sheet_obj.cell(row = i, column = 1)
                label_obj = sheet_obj.cell(row = i, column = 2)
                gw_obj = sheet_obj.cell(row = i, column = 3)
                floor_obj = sheet_obj.cell(row = i, column = 4)
                room_obj = sheet_obj.cell(row = i, column = 5)

                floor,room,type,count = str(label_obj.value).split('-')
                floor = floor.replace('O','')
                floor = floor.replace('E','')
                label_short = room + "-" + count 
                label = str(label_obj.value)
                gateway = str(gw_obj.value)

                #log.info("eshellid: " + str(eshellid_obj.value) + " | label: " + str(label_obj.value) + " | gw: " + str(gw_obj.value) + 
                #    " | floor: " + str(floor_obj.value) + " | room: " + str(room_obj.value) + " | id: " + str(id_obj.value) ) 

                log.debug("label: " +  " | gw: " + str(gw_obj.value) + " | floor: " + str(floor_obj.value) + " | room: " + str(room_obj.value) + " | id: " + str(id_obj.value) ) 
                log.debug(str(floor))                                           
                log.debug(str(label_short))                                           
        except:
            pass  
    #print(sheet_obj)      


    return SensorData(deviceUuid, deviceId, siteUuid, type, label, label_short, gateway, room, int(floor), int(index), int(value), timestamp)


def send_sensor_data_to_influxdb(sensor_data):
    log.debug(sensor_data)
    json_body = [
        {
            'measurement': 'motion',
            'tags': {
                'deviceUuid': sensor_data.deviceUuid,
                'deviceId': sensor_data.deviceId,
                'siteUuid': sensor_data.siteUuid,
                'type': sensor_data.type,
                'label': sensor_data.label,
                'label_short': sensor_data.label_short,
                'gateway': sensor_data.gateway,
                'room': sensor_data.room,
                'room': sensor_data.room,
                'floor': sensor_data.floor,
                'room': sensor_data.room
            },
            'fields': {                
                'index': sensor_data.index,
                'value': sensor_data.value,
                'timestamp': sensor_data.timestamp
            }
        }
    ]
    log.debug(json_body)
    log.debug(influxdb_client.write_points(json_body,time_precision="ms",protocol = "json"))

def on_message(client, userdata, msg):
    #log.debug("message received " + str(msg.payload.decode("utf-8")))
    #log.debug("message topic=" + msg.topic)
    """The callback for when a PUBLISH message is received from the server."""
    sensor_data = _parse_mqtt_message(msg.payload.decode('utf-8'))
    if sensor_data is not None:
         send_sensor_data_to_influxdb(sensor_data)

def on_log(client, userdata, level, buf):
    log.info("log: ",buf)

def on_connect(client, userdata, flags, rc):
    """ The callback for when the client receives a CONNACK response from the server."""
    log.info('Connected to MQTT Broker with result code ' + str(rc))
    client.subscribe(config['MQTT_TOPIC'])

def on_subscribe(client, userdata, mid, granted_qos):
    log.info('subscribed ' + str(rc))

def on_disconnect(client, userdata, rc):
    if rc != 0:
        log.debug("Disconnect with result code " + str(rc) )


def mqtt_loop():
    log.debug("mqtt_loop")
    log.debug(config['MQTT_CLIENT_ID'])
    log.debug(config['MQTT_TOPIC'])

    mqtt_client = mqtt.Client(config['MQTT_CLIENT_ID'] )
    mqtt_client.tls_set_context(context=None)
    mqtt_client.enable_logger(logger=log)
    mqtt_client.username_pw_set(config['MQTT_USER'], config['MQTT_PASSWORD'])
    mqtt_client.on_connect = on_connect
    mqtt_client.on_message = on_message
    mqtt_client.on_disconnect = on_disconnect
    
    mqtt_client.on_log=on_log
    mqtt_client.connect(config['MQTT_ADDRESS'], int(config['MQTT_PORT']) )
    mqtt_client.subscribe(config['MQTT_TOPIC'],1)
    mqtt_client.loop_forever()

def runExcelReader(xlsFile, sheet="Vossloh und Schwabe Multisensor"):  

    if not os.path.exists(xlsFile):
        log.error("[bold red blink]Excel File not found " + xlsFile + " [/]", extra={"markup": True})          
        log.error("Exit")  

    try:
        wb_obj = openpyxl.load_workbook(xlsFile)
        global sheet_obj
        sheet_obj = wb_obj[sheet]        
    except:
        log.error("cant read excel file %s!", args.config)         
      
    return sheet_obj       

    
def runCSVReader(CSVFile):  
    # sourcery skip: inline-immediately-returned-variable
   
    if not os.path.exists(CSVFile):
        log.error("[bold red blink]Excel File not found " + CSVFile + " [/]", extra={"markup": True})          
        log.error("Exit")  

    global csvdata 
    csvdata = util.readCSV_into_Dict(CSVFile)

    return csvdata       
    # for row in csvdata:
    #      print(row)


def searchfile(path,fileextension):
    #Reads all files from Folder          
    for filename in os.listdir(path):
        if filename.endswith(fileextension) :        
            log.info("Process File: " + os.path.join(path, filename))
            #runCSVReader(os.path.join(path,filename))
            return os.path.join(path,filename)
 

def get_parser():
    """Get parser object """
    from argparse import ArgumentParser, ArgumentDefaultsHelpFormatter
    parser = argparse.ArgumentParser(description='Understand functioning')
    parser.add_argument("-l", "--logfile", type=str, required=False, help="Logfile", default="")
    parser.add_argument("-c", "--config", type=str, required=False, help="Config", default="")
    parser.add_argument("-x", "--xlsfolder", type=str, required=False,  metavar="FILE", help="Path to Excel Files Folder", default="") 
    parser.add_argument("-m", "--mway_path", type=str, required=False,  metavar="FILE", help="Path mway Thing-it CSV Files Folder", default="") 
    parser.add_argument("--verbose", "-v", dest="log_level", action="append_const", const=-1,)
    parser.add_argument("--quiet", "-q", dest="log_level", action="append_const", const=1,)
    return parser

def setLogging(args):   
    log_level = LOG_LEVELS.index(DEFAULT_LOG_LEVEL)
    for adjustment in args.log_level or ():
            log_level = min(len(LOG_LEVELS) - 1, max(log_level + adjustment, 0))

    log_level_name = LOG_LEVELS[log_level]

    global log
    log = logging.getLogger(__name__)    
    shell_handler = RichHandler()    
    log.setLevel(log_level_name)
    shell_handler.setLevel(log_level_name)    
    fmt_shell = '%(message)s'    
    shell_formatter = logging.Formatter(fmt_shell)
    shell_handler.setFormatter(shell_formatter)
    log.addHandler(shell_handler)
    
    if len(args.logfile) > 0:
        file_handler = logging.FileHandler(args.logfile)
        file_handler.setLevel(logging.DEBUG)
        fmt_file = '%(levelname)s %(asctime)s [%(filename)s:%(funcName)s:%(lineno)d] %(message)s'
        file_formatter = logging.Formatter(fmt_file)
        file_handler.setFormatter(file_formatter)
        log.addHandler(file_handler)        

def main():
    log.info("[bold green]MQTT2Influx [Mway] © Stefan Knaak - E-Shelter Security.  [/]", extra={"markup": True}) 
    
    #Folders
    cwd = os.getcwd()
    xls_path = export.checkInputFolderPath(args.xlsfolder, "xlsx")
    mway_path = export.checkInputFolderPath(args.mway_path, "mway_csv")

    log.debug("xls_path: " + str(xls_path))    
    log.debug("mway_csv: " + str(mway_path))    
    log.debug(config)    
    
    log.info("[bold cyan]Folder Mode[/]", extra={"markup": True}) 

    if not os.path.exists(xls_path):
        log.error("[bold red blink]xlsx Folder not found " + xls_path + " [/]", extra={"markup": True})  
        log.error("Exit")  
    else:  
        log.info("[bold cyan]Process Excel Files from " + xls_path + " [/]", extra={"markup": True})     
        assetlist = searchfile(xls_path,'xlsx')  
        global assetlist_data
        assetlist_data = runExcelReader(assetlist)      

    if not os.path.exists(mway_path):
        log.error("[bold red blink]deuta_csv Folder not found " + mway_path + " [/]", extra={"markup": True})  
        log.error("Exit")  
    else:  
        log.info("[bold cyan]Process deuta_csv Files from " + mway_path + " [/]", extra={"markup": True})     
        mway_csvFile = searchfile(mway_path,'csv') 
        global mway_CSV_Data
        mway_CSV_Data = runCSVReader(mway_csvFile)
    

    #Main Loop
    mqtt_loop()        

def config_read():
    config_f = configparser.ConfigParser()

    try:
        config_f.read_file(open(args.config))
    
        config['MQTT_ADDRESS']      = config_f.get('mqtt', 'MQTT_ADDRESS')
        config['MQTT_PORT']         = config_f.get('mqtt', 'MQTT_PORT')
        config['MQTT_USER']         = config_f.get('mqtt', 'MQTT_USER')
        config['MQTT_PASSWORD']     = config_f.get('mqtt', 'MQTT_PASSWORD')
        config['MQTT_TOPIC']        = config_f.get('mqtt', 'MQTT_TOPIC')
        config['MQTT_CLIENT_ID']    = config_f.get('mqtt', 'MQTT_CLIENT_ID')

        config['INFLUXDB_ADDRESS']  = config_f.get('influx', 'INFLUXDB_ADDRESS')
        config['INFLUXDB_PORT']     = config_f.get('influx', 'INFLUXDB_PORT')
        config['INFLUXDB_USER']     = config_f.get('influx', 'INFLUXDB_USER')
        config['INFLUXDB_PASSWORD'] = config_f.get('influx', 'INFLUXDB_PASSWORD')
        config['INFLUXDB_DATABASE'] = config_f.get('influx', 'INFLUXDB_DATABASE')
        config['INFLUXDB_USESLL']   = config_f.get('influx', 'INFLUXDB_USESLL')
        config['INFLUXDB_NOVERIFY'] = config_f.get('influx', 'INFLUXDB_NOVERIFY')
        config['INFLUXDB_CREATE']   = config_f.get('influx', 'INFLUXDB_CREATE')
    
    except configparser.ParsingError:
        log.error("Unable to parse config from file %s!", args.config)
        sys.exit(1)

              
if __name__ == "__main__":    

    #parse arguments
    args = get_parser().parse_args()

    #get config
    if not args.config:
        args.config = os.path.dirname(os.path.abspath(__file__)) + '/config_mway.conf'

    #get config data
    config = {}
    config_read()

    setLogging(args)

    #create influx connection
    influxdb_client = utils_influx_Client.influxClient(config,log)

    try:
        main()
    except BaseException:
        import sys
        print(sys.exc_info()[0])
        import traceback
        print(traceback.format_exc())
    # finally:
    #     print("Press Enter to continue ...")
    #     input()
